import os
from openpyxl import Workbook


def generate_test_data():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(project_root, "data")
    os.makedirs(data_dir, exist_ok=True)

    file_path = os.path.join(data_dir, "test_data.xlsx")
    wb = Workbook()

    # --- Login sheet ---
    ws1 = wb.active
    ws1.title = "LoginTestData"
    ws1.append(
        [
            "TestCaseID",
            "Email",
            "Password",
            "ExpectedMessage",
            "Result",
            "Screenshot",
            "Video",
        ]
    )
    ws1.append(
        [
            "DangNhap-1",
            "luuthianduyen@gmail.com",
            "luuthianduyen247",
            "Tài khoản của tôi",
            "",
            "",
            "",
        ]
    )
    ws1.append(
        [
            "DangNhap-2",
            "",
            "123456789",
            "Email là bắt buộc",
            "",
            "",
            "",
        ]
    )
    ws1.append(
        [
            "DangNhap-3",
            "luuthianduyen@gmail.com",
            "",
            "Mật khẩu là bắt buộc",
            "",
            "",
            "",
        ]
    )
    ws1.append(
        [
            "DangNhap-4",
            "sai_email@gmail.com",
            "sai_pass",
            "Sai tên đăng nhập hoặc mật khẩu",
            "",
            "",
            "",
        ]
    )

    # --- RegisterTutor sheet ---
    ws2 = wb.create_sheet("RegisterTutorTestData")
    ws2.append(
        [
            "TestCaseID",
            "Gender",
            "Name",
            "Phone",
            "Address",
            "Note",
            "ExpectedMessage",
            "Result",
            "Screenshot",
            "Video",
        ]
    )
    ws2.append(
        [
            "Hiring-1",
            "Anh",
            "Nguyen Van A",
            "0912345678",
            "Hà Nội",
            "",
            "Đăng ký thành công",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-2",
            "Anh",
            "",
            "0912345678",
            "Hà Nội",
            "",
            "Họ và tên là bắt buộc",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-3",
            "Anh",
            "Nguyen Van A",
            "",
            "Hà Nội",
            "",
            "Số điện thoại là bắt buộc",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-4",
            "Anh",
            "Nguyen Van A",
            "0912345678",
            "",
            "",
            "Hệ thống không báo lỗi và tiếp tục đăng ký",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-5",
            "Anh",
            "Nguyen Van A",
            "0912345678",
            "Hà Nội",
            "",
            "Hệ thống không báo lỗi và tiếp tục đăng ký",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-6",
            "Anh",
            "Nguyen Van A",
            "abc12345",
            "Hà Nội",
            "",
            "Hệ thống không báo lỗi",
            "",
            "",
            "",
        ]
    )
    ws2.append(
        [
            "Hiring-7",
            "Anh",
            "Nguyen Van A",
            "12345",
            "Hà Nội",
            "",
            "Số điện thoại ít nhất 10 chữ số",
            "",
            "",
            "",
        ]
    )

    # --- SearchTestData sheet ---
    ws3 = wb.create_sheet("SearchTestData")
    ws3.append(
        [
            "TestCaseID",
            "Keyword",
            "ExpectedCount",
            "ExpectedMessage",
            "Result",
            "Screenshot",
            "Video",
        ]
    )
    ws3.append(
        ["Search-1", "Lớp 10", 9, "Hiển thị đúng và đủ kết quả phù hợp", "", "", ""]
    )
    ws3.append(
        [
            "Search-2",
            "abcdefxyz",
            0,
            "Hiển thị không có lớp gia sư nào phù hợp",
            "",
            "",
            "",
        ]
    )
    ws3.append(["Search-3", "", 0, "Không hiển thị kết quả tìm kiếm nào", "", "", ""])
    ws3.append(
        [
            "Search-4",
            "@@@!!!",
            0,
            "Hiển thị không có lớp gia sư nào phù hợp",
            "",
            "",
            "",
        ]
    )
    ws3.append(
        ["Search-5", "lớp 10", 9, "Hiển thị đúng và đủ kết quả phù hợp", "", "", ""]
    )
    ws3.append(
        [
            "Search-6",
            "   Lớp 10   ",
            9,
            "Hiển thị đúng và đủ kết quả phù hợp",
            "",
            "",
            "",
        ]
    )
    ws3.append(
        ["Search-7", "10", None, "Hiển thị đúng và đủ kết quả phù hợp", "", "", ""]
    )
    ws3.append(
        ["Search-8", "   ", 0, "Không hiển thị kết quả tìm kiếm nào", "", "", ""]
    )

    wb.save(file_path)
    print(f"File test data đã được tạo tại: {file_path}")


if __name__ == "__main__":
    generate_test_data()
