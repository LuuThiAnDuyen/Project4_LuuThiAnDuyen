
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")


def _header_map(ws):
    """
    Trả về dict {header_name: column_index}.
    Yêu cầu hàng 1 là header: TestCaseID, Result, Screenshot, Video ...
    """
    mapping = {}
    for cell in ws[1]:
        if cell.value:
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def _find_row_by_tcid(ws, tcid, tc_col):
    """
    Tìm dòng có TestCaseID == tcid (so sánh kiểu string, strip()).
    Nếu không thấy thì tạo dòng mới ở cuối bảng.
    """
    tcid = str(tcid).strip()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=tc_col).value
        if val is not None and str(val).strip() == tcid:
            return r
    # Không thấy -> thêm dòng mới
    r = ws.max_row + 1
    ws.cell(row=r, column=tc_col, value=tcid)
    return r


def _set_hyperlink(cell, path):
    """
    Ghi hyperlink nếu file tồn tại; nếu không tồn tại thì vẫn ghi đường dẫn dạng text.
    """
    if not path:
        return
    # Chuẩn hoá dấu gạch chéo để Excel hiểu hyperlink trên Windows
    path = path.replace("\\", "/")
    cell.value = os.path.basename(path)
    cell.hyperlink = path
    cell.style = "Hyperlink"


def write_test_result(
    data_file: str,
    sheet_name: str,
    tcid: str,
    result: str,
    screenshot_path: str = None,
    video_path: str = None,
):
    """
    Ghi PASS/FAIL + hyperlink Screenshot/Video vào đúng hàng của TestCaseID.
    - Không nổ nếu thiếu cột; chỉ ghi những cột header có thật.
    - Tô màu ô Result (xanh PASS, đỏ FAIL).
    """
    if not os.path.exists(data_file):
        raise FileNotFoundError(f"Excel not found: {data_file}")

    try:
        wb = load_workbook(data_file)
    except PermissionError:
        # Excel đang mở sẽ khoá file: thông báo rõ ràng
        raise PermissionError(
            f"Không thể mở '{data_file}'. Hãy đóng Excel đang mở file này rồi chạy lại."
        )

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Tạo header tối thiểu
        ws.append(["TestCaseID", "Result", "Screenshot", "Video"])
    else:
        ws = wb[sheet_name]

    headers = _header_map(ws)
    tc_col = headers.get("TestCaseID", 1)
    row = _find_row_by_tcid(ws, tcid, tc_col)

    # Ghi Result
    res_col = headers.get("Result")
    if res_col:
        c = ws.cell(row=row, column=res_col)
        c.value = result
        if isinstance(result, str) and result.upper() == "PASS":
            c.fill = GREEN_FILL
            c.font = GREEN_FONT
        elif isinstance(result, str) and result.upper() == "FAIL":
            c.fill = RED_FILL
            c.font = RED_FONT

    # Ghi Screenshot (hyperlink)
    sc_col = headers.get("Screenshot")
    if sc_col and screenshot_path:
        _set_hyperlink(ws.cell(row=row, column=sc_col), screenshot_path)

    # Ghi Video (hyperlink)
    vd_col = headers.get("Video")
    if vd_col and video_path:
        _set_hyperlink(ws.cell(row=row, column=vd_col), video_path)

    wb.save(data_file)
    wb.close()
