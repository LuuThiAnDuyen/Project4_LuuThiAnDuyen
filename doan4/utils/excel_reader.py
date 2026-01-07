
import os
from typing import List, Tuple, Dict, Any, Optional
from openpyxl import load_workbook


def _project_root() -> str:
    # thư mục gốc repo (thư mục chứa folder 'utils')
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# =========================
# GIỮ NGUYÊN: Login data
# =========================
def read_login_data() -> List[Dict[str, Any]]:
    """
    Đọc file data/login_test_data.xlsx (sheet active) -> list[dict]
    Keys: tc, username, password, expected, note
    """
    project_root = _project_root()
    file_path = os.path.join(project_root, "data", "login_test_data.xlsx")
    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        rows: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:  # skip header
                continue
            rows.append(
                {
                    "tc": row[0],
                    "username": row[1],
                    "password": row[2],
                    "expected": row[3],
                    "note": row[4] if len(row) > 4 else None,
                }
            )
        return rows
    finally:
        wb.close()


# ===================================
# HÀM CHUNG: đọc sheet theo tuple
# ===================================
def read_test_data(file_path, sheet_name):
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # bỏ header
        # Bỏ qua dòng trống hoàn toàn
        if not any(row):
            continue
        data.append(row)
    wb.close()
    return data


# ==========================================
# HÀM CHUNG: đọc sheet theo header -> dict
# ==========================================
def read_sheet_as_dicts(file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    """
    Đọc một sheet và trả về list[dict] dựa trên header hàng 1.
    Header sẽ được chuẩn hoá: strip, lower, thay khoảng trắng bằng underscore.
    """
    if not os.path.isabs(file_path):
        file_path = os.path.join(_project_root(), file_path)

    wb = load_workbook(file_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại trong '{file_path}'")
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []

        # Chuẩn hoá header
        norm_headers: List[str] = []
        for h in headers:
            h = (h or "").strip()
            h = h.replace("\n", " ").replace("\r", " ")
            h = "_".join(h.split())
            norm_headers.append(h.lower())

        results: List[Dict[str, Any]] = []
        for row in rows_iter:
            item: Dict[str, Any] = {}
            for idx, key in enumerate(norm_headers):
                val = row[idx] if idx < len(row) else None
                item[key] = val
            results.append(item)
        return results
    finally:
        wb.close()


# ======================================================
# TIỆN ÍCH: đọc SearchTestData → list[dict] chuẩn khoá
# ======================================================
def read_search_data(
    file_name: str = "test_data.xlsx",
    sheet_name: str = "SearchTestData",
) -> List[Dict[str, Any]]:
    """
    Đọc data/test_data.xlsx (mặc định) sheet SearchTestData.
    Trả về list[dict] với các key thống nhất:
      - tcid, keyword, expected_count, expected_message, result, screenshot, video

    Hỗ trợ 2 kiểu sheet:
      a) Có header đúng tên cột (tcid/keyword/expected_count/expected_message/result/screenshot/video)
      b) Không header, data theo thứ tự 7 cột => map theo index.
    """
    file_path = os.path.join(_project_root(), "data", file_name)

    # Thử đọc theo header trước
    try:
        dict_rows = read_sheet_as_dicts(file_path, sheet_name)
        if dict_rows:
            out: List[Dict[str, Any]] = []
            for r in dict_rows:
                out.append(
                    {
                        "tcid": r.get("tcid") or r.get("id") or r.get("tc") or "",
                        "keyword": r.get("keyword"),
                        "expected_count": r.get("expected_count"),
                        "expected_message": r.get("expected_message"),
                        "result": r.get("result"),
                        "screenshot": r.get("screenshot"),
                        "video": r.get("video"),
                    }
                )
            # Nếu row nào không có tcid thì loại bỏ
            out = [x for x in out if str(x.get("tcid", "")).strip()]
            if out:
                return out
    except Exception:
        # rơi xuống cách đọc dạng tuple
        pass

    # Đọc dạng tuple (không header)
    tuple_rows = read_test_data(file_path, sheet_name)
    out2: List[Dict[str, Any]] = []
    for row in tuple_rows:
        # kỳ vọng: (tcid, keyword, expected_count, expected_message, result, screenshot, video)
        tcid = (row[0] if len(row) > 0 else "") or ""
        keyword = row[1] if len(row) > 1 else None
        expected_count = row[2] if len(row) > 2 else None
        expected_message = row[3] if len(row) > 3 else None
        result = row[4] if len(row) > 4 else None
        screenshot = row[5] if len(row) > 5 else None
        video = row[6] if len(row) > 6 else None

        if str(tcid).strip():
            out2.append(
                {
                    "tcid": tcid,
                    "keyword": keyword,
                    "expected_count": expected_count,
                    "expected_message": expected_message,
                    "result": result,
                    "screenshot": screenshot,
                    "video": video,
                }
            )
    return out2
