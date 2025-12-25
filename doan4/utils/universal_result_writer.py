# utils/universal_result_writer.py
from __future__ import annotations

import os
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


# =========================
# Common helpers
# =========================
def _project_root() -> str:
    # thư mục gốc project (chứa folder 'utils')
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _abs_path(path: str) -> str:
    return path if os.path.isabs(path) else os.path.join(_project_root(), path)


def _norm_key(k: str) -> str:
    k = (k or "").strip()
    k = k.replace("\n", " ").replace("\r", " ")
    k = "_".join(k.split())
    return k.lower()


def _coalesce(d: Dict[str, Any], keys: Sequence[str], default: Any = "") -> Any:
    for k in keys:
        if k in d and d.get(k) not in (None, ""):
            return d.get(k)
    return default


def _get_tcid(row: Dict[str, Any]) -> str:
    return str(
        _coalesce(row, ["testcaseid", "test_case_id", "tcid", "id", "tc"], default="")
    ).strip()


def _ensure_parent_dir(file_path: str) -> None:
    p = Path(file_path)
    if p.parent and not p.parent.exists():
        p.parent.mkdir(parents=True, exist_ok=True)


# =========================
# CSV
# =========================
def _read_csv_rows(
    file_path: str, encoding: str = "utf-8"
) -> Tuple[List[str], List[Dict[str, Any]]]:
    with open(file_path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows: List[Dict[str, Any]] = []
        for r in reader:
            rows.append({_norm_key(k): v for k, v in (r or {}).items()})
    # normalize fieldnames as-is for write back (keep original order)
    return fieldnames, rows


def _write_csv_rows(
    file_path: str,
    fieldnames_original: List[str],
    rows_norm: List[Dict[str, Any]],
    encoding: str = "utf-8",
) -> None:
    # write with original headers if possible, otherwise use normalized keys sorted
    if fieldnames_original:
        out_headers = fieldnames_original
        # when writing, map normalized keys back to original headers by normalization
        header_map = {_norm_key(h): h for h in out_headers}
        with open(file_path, "w", encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=out_headers)
            writer.writeheader()
            for r in rows_norm:
                out_row = {}
                for nk, v in r.items():
                    ok = header_map.get(nk)
                    if ok:
                        out_row[ok] = v
                # ensure all headers exist (avoid KeyError)
                for h in out_headers:
                    out_row.setdefault(h, "")
                writer.writerow(out_row)
        return

    # fallback (rare)
    out_headers = sorted({k for r in rows_norm for k in r.keys()})
    with open(file_path, "w", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=out_headers)
        writer.writeheader()
        for r in rows_norm:
            writer.writerow(r)


def _write_test_result_to_csv(
    file_path: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    encoding: str = "utf-8",
) -> None:
    headers, rows = _read_csv_rows(file_path, encoding=encoding)

    updated = False
    for r in rows:
        if _get_tcid(r) == str(tcid).strip():
            r["result"] = result
            r["screenshot"] = screenshot
            r["video"] = video
            updated = True
            break

    if not updated:
        # append new row using normalized keys
        rows.append(
            {
                "testcaseid": str(tcid).strip(),
                "email": "",
                "password": "",
                "expectedmessage": "",
                "result": result,
                "screenshot": screenshot,
                "video": video,
            }
        )

    _write_csv_rows(file_path, headers, rows, encoding=encoding)


# =========================
# JSON
# =========================
def _read_json_any(file_path: str, encoding: str = "utf-8") -> Any:
    return json.loads(Path(file_path).read_text(encoding=encoding))


def _write_json_any(file_path: str, data: Any, encoding: str = "utf-8") -> None:
    _ensure_parent_dir(file_path)
    Path(file_path).write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding=encoding
    )


def _write_test_result_to_json(
    file_path: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    encoding: str = "utf-8",
) -> None:
    data = _read_json_any(file_path, encoding=encoding)

    def _norm_obj(obj: Dict[str, Any]) -> Dict[str, Any]:
        return {_norm_key(k): v for k, v in (obj or {}).items()}

    # Determine container
    if isinstance(data, list):
        arr = data
        container_type = "list"
    elif isinstance(data, dict) and isinstance(data.get("cases"), list):
        arr = data["cases"]
        container_type = "dict_cases"
    elif isinstance(data, dict):
        # single case dict
        arr = [data]
        container_type = "single_dict"
    else:
        raise ValueError("Unsupported JSON structure for login test data")

    updated = False
    for i, obj in enumerate(arr):
        if not isinstance(obj, dict):
            continue
        norm = _norm_obj(obj)
        if _get_tcid(norm) == str(tcid).strip():
            # update in original object keys if present, else add canonical keys
            obj["Result"] = result
            obj["Screenshot"] = screenshot
            obj["Video"] = video
            updated = True
            break

    if not updated:
        new_obj = {
            "TestCaseID": str(tcid).strip(),
            "Email": "",
            "Password": "",
            "ExpectedMessage": "",
            "Result": result,
            "Screenshot": screenshot,
            "Video": video,
        }
        arr.append(new_obj)

    # Re-attach for single_dict mode
    if container_type == "single_dict":
        data = arr[0]
    _write_json_any(file_path, data, encoding=encoding)


# =========================
# Excel
# =========================
def _write_test_result_to_excel(
    file_path: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    sheet_name: str = "LoginTestData",
) -> None:
    wb = load_workbook(file_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại trong '{file_path}'")
        ws = wb[sheet_name]

        # read header row (assume first row)
        headers = [str(c.value or "").strip() for c in ws[1]]
        headers_norm = [_norm_key(h) for h in headers]

        def _col_index(key: str) -> Optional[int]:
            key_n = _norm_key(key)
            for idx, h in enumerate(headers_norm, start=1):
                if h == key_n:
                    return idx
            return None

        idx_tcid = _col_index("TestCaseID") or _col_index("tcid") or _col_index("id")
        idx_result = _col_index("Result")
        idx_shot = _col_index("Screenshot")
        idx_vid = _col_index("Video")

        if not idx_tcid:
            raise ValueError("Excel thiếu cột TestCaseID")

        # If missing result/screenshot/video columns, create them at end
        max_col = ws.max_column

        def _ensure_col(name: str) -> int:
            nonlocal max_col
            idx = _col_index(name)
            if idx:
                return idx
            max_col += 1
            ws.cell(row=1, column=max_col).value = name
            headers_norm.append(_norm_key(name))
            return max_col

        idx_result = idx_result or _ensure_col("Result")
        idx_shot = idx_shot or _ensure_col("Screenshot")
        idx_vid = idx_vid or _ensure_col("Video")

        target_row = None
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=idx_tcid).value
            if str(val or "").strip() == str(tcid).strip():
                target_row = r
                break

        if target_row is None:
            # append new row
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=idx_tcid).value = str(tcid).strip()

        ws.cell(row=target_row, column=idx_result).value = result
        ws.cell(row=target_row, column=idx_shot).value = screenshot
        ws.cell(row=target_row, column=idx_vid).value = video

        wb.save(file_path)
    finally:
        wb.close()


# =========================
# YAML
# =========================
def _read_yaml_any(file_path: str, encoding: str = "utf-8") -> Any:
    try:
        import yaml  # pip install pyyaml
    except ImportError as e:
        raise ImportError("Thiếu PyYAML. Cài: pip install pyyaml") from e
    return yaml.safe_load(Path(file_path).read_text(encoding=encoding))


def _write_yaml_any(file_path: str, data: Any, encoding: str = "utf-8") -> None:
    try:
        import yaml
    except ImportError as e:
        raise ImportError("Thiếu PyYAML. Cài: pip install pyyaml") from e
    _ensure_parent_dir(file_path)
    # sort_keys=False để giữ thứ tự “nhìn dễ”
    text = yaml.safe_dump(data, allow_unicode=True, sort_keys=False)
    Path(file_path).write_text(text, encoding=encoding)


def _write_test_result_to_yaml(
    file_path: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    encoding: str = "utf-8",
) -> None:
    data = _read_yaml_any(file_path, encoding=encoding)

    def _norm_obj(obj: Dict[str, Any]) -> Dict[str, Any]:
        return {_norm_key(k): v for k, v in (obj or {}).items()}

    if isinstance(data, list):
        arr = data
        container_type = "list"
    elif isinstance(data, dict) and isinstance(data.get("cases"), list):
        arr = data["cases"]
        container_type = "dict_cases"
    elif isinstance(data, dict):
        arr = [data]
        container_type = "single_dict"
    else:
        raise ValueError("Unsupported YAML structure for login test data")

    updated = False
    for obj in arr:
        if not isinstance(obj, dict):
            continue
        norm = _norm_obj(obj)
        if _get_tcid(norm) == str(tcid).strip():
            # write using friendly keys
            obj["Result"] = result
            obj["Screenshot"] = screenshot
            obj["Video"] = video
            updated = True
            break

    if not updated:
        arr.append(
            {
                "TestCaseID": str(tcid).strip(),
                "Email": "",
                "Password": "",
                "ExpectedMessage": "",
                "Result": result,
                "Screenshot": screenshot,
                "Video": video,
            }
        )

    if container_type == "single_dict":
        data = arr[0]

    _write_yaml_any(file_path, data, encoding=encoding)


# =========================
# XML
# =========================
def _xml_local(tag: str) -> str:
    return tag.split("}")[-1] if tag else tag


def _find_or_create_child(parent, tag_name: str):
    import xml.etree.ElementTree as ET

    for ch in list(parent):
        if _xml_local(ch.tag).lower() == tag_name.lower():
            return ch
    return ET.SubElement(parent, tag_name)


def _write_test_result_to_xml(
    file_path: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    encoding: str = "utf-8",
) -> None:
    import xml.etree.ElementTree as ET

    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"XML file not found: {file_path}")

    try:
        tree = ET.parse(str(p))
        root = tree.getroot()
    except ET.ParseError:
        root = ET.fromstring(p.read_text(encoding=encoding))
        tree = ET.ElementTree(root)

    testcases = root.findall(".//TestCase")
    target = None

    for tc in testcases:
        tcid_node = None
        for ch in list(tc):
            if _xml_local(ch.tag).lower() in (
                "testcaseid",
                "test_case_id",
                "tcid",
                "id",
            ):
                tcid_node = ch
                break
        if (
            tcid_node is not None
            and (tcid_node.text or "").strip() == str(tcid).strip()
        ):
            target = tc
            break

    if target is None:
        # create new TestCase
        target = ET.SubElement(root, "TestCase")
        _find_or_create_child(target, "TestCaseID").text = str(tcid).strip()

    _find_or_create_child(target, "Result").text = str(result or "")
    _find_or_create_child(target, "Screenshot").text = str(screenshot or "")
    _find_or_create_child(target, "Video").text = str(video or "")

    try:
        ET.indent(tree, space="  ", level=0)  # Python 3.9+
    except Exception:
        pass

    tree.write(str(p), encoding=encoding, xml_declaration=True)


# =========================
# Public API
# =========================
def write_test_result_any(
    source: str,
    tcid: str,
    result: str,
    screenshot: str = "",
    video: str = "",
    sheet_name: str = "LoginTestData",
    encoding: str = "utf-8",
) -> None:
    """
    Ghi/update Result + Screenshot + Video theo TestCaseID vào file dữ liệu.

    Hỗ trợ:
      - .xlsx/.xls  (Excel)
      - .csv
      - .json
      - .txt        (key=value blocks separated by ---)
      - .yml/.yaml
      - .xml
    """
    source_abs = _abs_path(source)
    ext = Path(source_abs).suffix.lower()

    if ext in [".xlsx", ".xls"]:
        _write_test_result_to_excel(
            file_path=source_abs,
            tcid=tcid,
            result=result,
            screenshot=screenshot,
            video=video,
            sheet_name=sheet_name,
        )
        return

    if ext == ".csv":
        _write_test_result_to_csv(
            file_path=source_abs,
            tcid=tcid,
            result=result,
            screenshot=screenshot,
            video=video,
            encoding=encoding,
        )
        return

    if ext == ".json":
        _write_test_result_to_json(
            file_path=source_abs,
            tcid=tcid,
            result=result,
            screenshot=screenshot,
            video=video,
            encoding=encoding,
        )
        return

    if ext in [".yml", ".yaml"]:
        _write_test_result_to_yaml(
            file_path=source_abs,
            tcid=tcid,
            result=result,
            screenshot=screenshot,
            video=video,
            encoding=encoding,
        )
        return

    if ext == ".xml":
        _write_test_result_to_xml(
            file_path=source_abs,
            tcid=tcid,
            result=result,
            screenshot=screenshot,
            video=video,
            encoding=encoding,
        )
        return

    raise ValueError(f"Unsupported format for write_test_result_any: {ext}")


def write_test_results_bulk_any(
    source: str,
    results: Sequence[Dict[str, str]],
    sheet_name: str = "LoginTestData",
    encoding: str = "utf-8",
) -> None:
    """
    Bulk writer tiện dụng:
    results: list of dict, mỗi dict có keys:
      - tcid (or TestCaseID)
      - result
      - screenshot (optional)
      - video (optional)

    Gọi internal writer theo từng testcase.
    """
    for r in results:
        tcid = r.get("tcid") or r.get("TestCaseID") or r.get("testcaseid") or ""
        if not str(tcid).strip():
            continue
        write_test_result_any(
            source=source,
            tcid=str(tcid).strip(),
            result=r.get("result", ""),
            screenshot=r.get("screenshot", ""),
            video=r.get("video", ""),
            sheet_name=sheet_name,
            encoding=encoding,
        )
