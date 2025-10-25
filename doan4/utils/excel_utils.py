import os
from openpyxl import Workbook, load_workbook

SHEET_NAME = "LoginTestData"
HEADERS = ["Test Case", "Username/Email", "Password", "Expected Result", "Note"]


def get_excel_path():
    project_root = os.path.dirname(os.path.abspath(os.path.join(__file__, "..")))
    data_dir = os.path.join(project_root, "data")
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, "login_test_data.xlsx")


def ensure_workbook(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(path)
        wb.close()
    else:
        # bảo đảm có sheet & header
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
            wb.save(path)
        wb.close()


def append_row(path: str, row: list):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    ws.append(row)
    wb.save(path)
    wb.close()


def read_rows(path: str):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        data.append(list(row))
    wb.close()
    return data


def next_tc_id(path: str) -> str:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    max_num = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        tcid = (row[0] or "").strip().upper()
        if tcid.startswith("TC"):
            num = tcid[2:]
            if num.isdigit():
                max_num = max(max_num, int(num))
    wb.close()
    return f"TC{max_num+1:02d}"
