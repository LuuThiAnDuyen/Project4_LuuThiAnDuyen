from __future__ import annotations

import os
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
import pytest

from openpyxl import load_workbook


def _project_root() -> str:
    # thư mục gốc project (chứa folder 'utils')
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _abs_path(path: str) -> str:
    return path if os.path.isabs(path) else os.path.join(_project_root(), path)


def _norm_key(k: str) -> str:
    k = (k or "").strip()
    k = k.replace("\n", " ").replace("\r", " ")
    k = "_".join(k.split())
    return k.lower()


# ========== Excel ==========
def _read_login_from_excel(
    file_path: str, sheet_name: str = "LoginTestData"
) -> List[Dict[str, Any]]:
    """
    Excel format (theo ảnh bạn gửi):
      TestCaseID | Email | Password | ExpectedMessage | Result | Screenshot | Video
    """
    wb = load_workbook(file_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại trong '{file_path}'")
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []

        headers_norm = [_norm_key(str(h or "")) for h in headers]
        out: List[Dict[str, Any]] = []

        for row in rows_iter:
            if not any(row):
                continue
            item: Dict[str, Any] = {}
            for i, key in enumerate(headers_norm):
                item[key] = row[i] if i < len(row) else None
            out.append(item)

        return out
    finally:
        wb.close()


# ========== CSV ==========
def _read_login_from_csv(
    file_path: str, encoding: str = "utf-8"
) -> List[Dict[str, Any]]:
    with open(file_path, encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        rows: List[Dict[str, Any]] = []
        for r in reader:
            item = {_norm_key(k): v for k, v in (r or {}).items()}
            rows.append(item)
        return rows


# ========== JSON ==========
def _read_login_from_json(file_path: str) -> List[Dict[str, Any]]:
    data = json.loads(Path(file_path).read_text(encoding="utf-8"))

    if isinstance(data, list):
        return [
            {_norm_key(k): v for k, v in obj.items()}
            for obj in data
            if isinstance(obj, dict)
        ]

    if isinstance(data, dict):
        if "cases" in data and isinstance(data["cases"], list):
            return [
                {_norm_key(k): v for k, v in obj.items()}
                for obj in data["cases"]
                if isinstance(obj, dict)
            ]
        return [{_norm_key(k): v for k, v in data.items()}]

    raise ValueError("JSON login data must be list or dict")


# ========== YAML ==========
def _read_login_from_yaml(
    file_path: str, encoding: str = "utf-8"
) -> List[Dict[str, Any]]:
    """
    YAML hỗ trợ:
    - list of dict:
        - TestCaseID: ...
          Email: ...
    - hoặc dict có key "cases": [ ... ]
    """
    try:
        import yaml  # pip install pyyaml
    except ImportError as e:
        raise ImportError("Thiếu thư viện PyYAML. Cài bằng: pip install pyyaml") from e

    text = Path(file_path).read_text(encoding=encoding)
    data = yaml.safe_load(text)

    def _norm_obj(obj: Dict[str, Any]) -> Dict[str, Any]:
        return {_norm_key(k): v for k, v in (obj or {}).items()}

    if isinstance(data, list):
        return [_norm_obj(x) for x in data if isinstance(x, dict)]

    if isinstance(data, dict):
        if "cases" in data and isinstance(data["cases"], list):
            return [_norm_obj(x) for x in data["cases"] if isinstance(x, dict)]
        # coi như single object
        return [_norm_obj(data)]

    raise ValueError("YAML login data must be list or dict")


# ========== XML ==========
def _read_login_from_xml(
    file_path: str, encoding: str = "utf-8"
) -> List[Dict[str, Any]]:
    """
    XML kỳ vọng dạng:
    <LoginTestData>
      <TestCase>
        <TestCaseID>...</TestCaseID>
        <Email>...</Email>
        ...
      </TestCase>
    </LoginTestData>

    Linh hoạt: nếu tag không đúng tên, vẫn đọc tất cả children dưới mỗi <TestCase>.
    """
    import xml.etree.ElementTree as ET

    # ET.parse tự xử lý encoding theo header; fallback: đọc text
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
    except ET.ParseError:
        content = Path(file_path).read_text(encoding=encoding)
        root = ET.fromstring(content)

    # Tìm các node "TestCase" (không phân biệt namespace đơn giản)
    cases = root.findall(".//TestCase")
    rows: List[Dict[str, Any]] = []

    for c in cases:
        item: Dict[str, Any] = {}
        for child in list(c):
            tag = child.tag.split("}")[-1]  # handle namespace
            item[_norm_key(tag)] = (child.text or "").strip()
        if item:
            rows.append(item)

    # Nếu không có <TestCase>, thử coi root chính là 1 testcase
    if not rows and list(root):
        item = {}
        for child in list(root):
            tag = child.tag.split("}")[-1]
            item[_norm_key(tag)] = (child.text or "").strip()
        if item:
            rows.append(item)

    return rows


def read_login_cases_any(
    source: str,
    sheet_name: str = "LoginTestData",
) -> List[Tuple[str, str, str, str]]:
    """
    Trả về ĐÚNG 4 field cho test:
    (tcid, email, password, expected_message)

    KHÔNG load: Result, Screenshot, Video
    """

    source_abs = _abs_path(source)
    ext = Path(source_abs).suffix.lower()

    # ===== xác định source =====
    if ext in [".xlsx", ".xls"]:
        rows = _read_login_from_excel(source_abs, sheet_name=sheet_name)
    elif ext == ".csv":
        rows = _read_login_from_csv(source_abs)
    elif ext == ".json":
        rows = _read_login_from_json(source_abs)
    elif ext in [".yml", ".yaml"]:
        rows = _read_login_from_yaml(source_abs)
    elif ext == ".xml":
        rows = _read_login_from_xml(source_abs)
    else:
        raise ValueError(f"Unsupported login data file type: {ext}")

    cases: List[Tuple[str, str, str, str]] = []

    for r in rows:
        tcid = (
            r.get("testcaseid")
            or r.get("test_case_id")
            or r.get("tcid")
            or r.get("id")
            or r.get("tc")
            or ""
        )

        email = r.get("email") or ""
        password = r.get("password") or ""

        expected_message = (
            r.get("expectedmessage")
            or r.get("expected_message")
            or r.get("expectedmes")
            or r.get("expected")
            or ""
        )

        if str(tcid).strip():
            cases.append(
                (
                    str(tcid).strip(),
                    str(email),
                    str(password),
                    str(expected_message),
                )
            )

    return cases
